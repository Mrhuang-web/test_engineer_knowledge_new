import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# 获取当前脚本所在目录
current_dir = os.path.dirname(os.path.abspath(__file__))
# 目标文件目录，相对路径（CSV文件在当前目录下的20260121子目录中）
files_dir = os.path.join(current_dir, '20260121')
# 输出文件路径
output_file = os.path.join(current_dir, 'file_analysis.xlsx')

def process_files():
    # 创建Excel工作簿
    wb = Workbook()
    
    # 设置样式
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # 处理Sheet1：统计文件行数
    ws1 = wb.active
    ws1.title = 'Sheet1'
    
    # 设置Sheet1表头
    ws1.append(['filename', 'number'])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 调整Sheet1列宽
    ws1.column_dimensions['A'].width = 50
    ws1.column_dimensions['B'].width = 15
    
    # 处理Sheet2：表头和第一行数据
    ws2 = wb.create_sheet(title='Sheet2')
    
    # 设置Sheet2表头
    ws2.append(['filename'])
    ws2.cell(row=1, column=2, value='字段')
    
    # 应用表头样式
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    # 调整Sheet2初始列宽
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 30
    
    # 获取所有CSV文件
    csv_files = [f for f in os.listdir(files_dir) if f.endswith('.csv')]
    
    # 遍历处理每个文件
    for filename in csv_files:
        file_path = os.path.join(files_dir, filename)
        print(f'Processing {filename}...')
        
        try:
            # 直接统计文件行数（表头不计入）
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                total_lines = sum(1 for _ in f)
            # 减去表头行
            row_count = max(0, total_lines - 1)
            
            # 写入行数统计到Sheet1
            ws1.append([filename, row_count])
            # 设置Sheet1数据行样式
            for cell in ws1[ws1.max_row]:
                cell.border = border
                if cell.column == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
            
            # 处理Sheet2
            # 读取文件内容，按^分裂处理
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                # 读取表头行
                header_line = f.readline().strip()
                # 读取第一行数据
                first_data_line = f.readline().strip()
                
                if header_line and first_data_line:
                    # 按^分裂表头字段
                    header_fields = header_line.split('^')
                    # 按^分裂数据字段
                    data_fields = first_data_line.split('^')
                    
                    # 计算当前文件的起始行
                    start_row = ws2.max_row + 1
                    
                    # 写入filename，占两行（第一列，上下两个单元格）
                    ws2.cell(row=start_row, column=1, value=filename)
                    ws2.cell(row=start_row+1, column=1, value='')
                    ws2.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
                    ws2.cell(row=start_row, column=1).font = Font(bold=True)
                    ws2.cell(row=start_row, column=1).border = border
                    ws2.cell(row=start_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 写入字段行
                    for col_idx, field_name in enumerate(header_fields, start=2):
                        cell = ws2.cell(row=start_row, column=col_idx, value=field_name)
                        cell.border = border
                        cell.alignment = left_align
                        cell.font = Font(bold=True)
                        
                        # 动态调整列宽
                        col_letter = cell.column_letter
                        if col_letter not in ws2.column_dimensions:
                            ws2.column_dimensions[col_letter].width = 30
                        current_width = ws2.column_dimensions[col_letter].width
                        new_width = max(current_width, len(str(field_name)) + 5)
                        ws2.column_dimensions[col_letter].width = min(new_width, 50)
                    
                    # 写入数据行
                    for col_idx, data_value in enumerate(data_fields, start=2):
                        cell = ws2.cell(row=start_row+1, column=col_idx, value=data_value)
                        cell.border = border
                        cell.alignment = left_align
                        
                        # 动态调整列宽
                        col_letter = cell.column_letter
                        if col_letter not in ws2.column_dimensions:
                            ws2.column_dimensions[col_letter].width = 30
                        current_width = ws2.column_dimensions[col_letter].width
                        new_width = max(current_width, len(str(data_value)) + 5)
                        ws2.column_dimensions[col_letter].width = min(new_width, 50)
        
        except Exception as e:
            print(f'Error processing {filename}: {e}')
            ws1.append([filename, f'Error: {str(e)}'])
            # 设置错误行样式
            for cell in ws1[ws1.max_row]:
                cell.border = border
                if cell.column == 1:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
    
    # 保存Excel文件
    wb.save(output_file)
    print(f'Processing completed. Results saved to {output_file}')

if __name__ == '__main__':
    process_files()