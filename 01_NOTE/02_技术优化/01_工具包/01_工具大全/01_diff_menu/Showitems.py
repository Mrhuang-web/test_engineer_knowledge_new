# -*- coding: utf-8 -*-
"""
@Time ： 2024/12/16 14:09
@Auth ： sunzhonghua
@File ：tt.py.py
@IDE ：PyCharm
"""
import json
import time
import os

import openpyxl
import pandas as pd
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import shutil


# 定义一个函数来复制单元格样式
def copy_cell_style(source, target):
    target.font = source.font
    target.border = source.border
    target.fill = source.fill
    target.number_format = source.number_format
    target.protection = source.protection
    target.alignment = source.alignment


def get_items(xlsx_name, rolejson):
    # 指定文件路径
    # file_path = 'path/to/your/textfile.txt'
    # 使用with语句打开文件，确保文件会被正确关闭
    with open(rolejson, 'r', encoding='utf-8') as file:
        # 读取文件内容
        content = file.read()
    # print(content)

    json_data = json.loads(content)
    file_name = xlsx_name
    # 创建一个空的DataFrame
    empty_df = pd.DataFrame(columns=['导航', '二级', '三级', '四级', '五级'])  # 你可以根据需要自定义列名
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        empty_df.to_excel(writer, index=False, sheet_name='Sheet1')
        pass
    print(f"文件 '{file_name}' 已创建。")

    # todo 递归函数来提取text值 [一个级的list]
    def extract_text(node, level=0, path=None, all_paths=None):
        # todo 每一层子层
        if path is None:
            path = []
        # todo 一个层级的所有子层
        if all_paths is None:
            all_paths = []

        # todo 主要就是靠这个判断来取即可【重点】
        if 'text' in node and 'id' in node and 'canEdit' in node:  # 控制是否选中
            # print("level=====", level)
            # print("前", path)
            # if path.__len__() == level:
            if node['checked'] is True:
                path.append(node['text'])
                path.append(node['id'])

            # print("后", path)
        # todo 这块是根据是否有children来判断是否继续递归 - 递归的判断也是前面的部分
        if 'children' in node and 'text' in node and 'id' in node and 'checked' in node:  # 控制是否选中
            if node['checked'] is True:
                for child in node['children']:
                    # print("children")
                    # todo 递归开始
                    extract_text(child, level + 1, path + [node['text']] + [node['id']], all_paths)
        else:
            all_paths.append(path)

        return all_paths

    # todo 提取text值
    node = (json_data['data']['treeNodes'])
    # todo 有多个menu级 - 因此需要用到for去对每个级进行子集处理
    for listone in node:
        # todo 递归调用，正式开始
        all_paths = extract_text(node=listone)
        # todo 层级获取不到只能返回空，会导致max错误、需要做个判断
        if len(all_paths) > 1:
            max_length = max(len(path) for path in all_paths)  # 找到最深的层级
            columns = [f'Level {i + 1}' for i in range(max_length - 1)]
            rows = [path + [None] * (max_length - len(path)) for path in all_paths]  # 确保所有路径长度相同
            df = pd.DataFrame(rows)
            # 将DataFrame追加写入Excel文件
            try:
                # 尝试加载已存在的工作簿
                book = load_workbook(file_name)
                startrow = book['Sheet1'].max_row
                book.close()
            except FileNotFoundError:
                # 如果文件不存在， startrow为0
                startrow = 0
            writer = pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            df.to_excel(writer, index=False, sheet_name='Sheet1', startrow=startrow, header=False)
            # 关闭写入器，这将保存文件
            writer.close()

    # 每行去掉重复的列
    wb = load_workbook(xlsx_name)
    ws = wb.active
    # 遍历每一行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):  # 假设第一行是标题行，从第二行开始处理
        i = 2  # 从第三列开始
        while i < len(row) - 1:  # 确保有足够的列进行检查
            # print(i)
            # 检查第i列和第i+1列是否分别与第0列和第1列相同
            if row[i].value is None:
                break
            if row[i].value == row[i - 2].value and row[i + 1].value == row[i - 1].value:
                for j in range(i, len(row) - 2):
                    row[j].value = row[j + 2].value
                # 由于前移了2列，将最后2列的数据设置为空
                row[-1].value = None
                row[-2].value = None
            else:
                i += 2  # 移动到下一对列进行检查
    # 保存修改后的Excel文件
    wb.save(xlsx_name)


# 备份文件
def backup_file(original, backup):
    shutil.copyfile(original, backup)


def timesign():
    from datetime import datetime
    # 获取当前时间
    now = datetime.now()
    # 按照指定格式输出时间
    formatted_time = now.strftime('%Y%m%d%H%M%S')
    print(formatted_time)
    return formatted_time


def pk_difference(file1="A.xlsx", file2="B.xlsx"):
    """
    python 对比两个excel文件A.xlsx(测试环境)和B.xlsx(现网),
    如果A的行在B中不存在,就将该行数据追加写到写到B表格的末尾且标记为红色；
    如果A的行在B中存在,就在B的对应行标记绿色
    #B表格中
    绿色：测试环境和现网都存在的会标记绿色---一般升级后应该是一致的
    白色：没有颜色标识的是测试环境没有但是现网确有的---测试环境缺的菜单或者按钮
    红色：测试环境有但是现网不存在的---应该是新增或者修改的菜单或者按钮

    """
    # 备份A.xlsx和B.xlsx到同一目录下的不同文件名
    file1_bak_name = './BakFile/' + file1.split('/')[2][:-5] + '_' + timesign() + '.xlsx'
    file2_bak_name = './BakFile/' + file2.split('/')[2][:-5] + '_' + timesign() + '.xlsx'
    backup_file(file1, file1_bak_name)
    backup_file(file2, file2_bak_name)
    # 打开A.xlsx和B.xlsx工作簿
    wb_a = openpyxl.load_workbook(file1)
    ws_a = wb_a.active
    wb_b = openpyxl.load_workbook(file2)
    ws_b = wb_b.active
    # 标记颜色
    green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    # 读取A.xlsx中的数据
    data_a = []
    for row in ws_a.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，从第二行开始读取
        data_a.append(row)
    # 读取B.xlsx中的数据
    data_b = []
    for row in ws_b.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题行，从第二行开始读取
        data_b.append(row)

    # 遍历A.xlsx中的每一行
    for row_a in data_a:
        # 检查行是否在B.xlsx中存在
        found = False
        for index, row_b in enumerate(data_b):
            if row_a == row_b:
                # 如果存在，在B.xlsx的对应行标记绿色
                for col_index, _ in enumerate(row_a):
                    ws_b.cell(row=index + 2, column=col_index + 1).fill = green_fill
                found = True
                break
        if not found:
            # 如果不存在，追加到B.xlsx的末尾
            new_row = ws_b.max_row + 1
            # new_row = len(ws_b.max_row) + 1
            ws_b.append(row_a)
            for col_index, _ in enumerate(row_a):
                ws_b.cell(row=new_row, column=col_index + 1).fill = red_fill
    # 保存B.xlsx工作簿
    wb_b.save(file2)
    # 关闭工作簿
    wb_a.close()
    wb_b.close()


if __name__ == '__main__':
    # pk_difference(file1="A.xlsx", file2="B.xlsx")
    # 对比内蒙

    pass
