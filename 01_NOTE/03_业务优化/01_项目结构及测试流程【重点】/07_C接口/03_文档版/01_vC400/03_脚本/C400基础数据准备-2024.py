#-*- coding:utf-8 -*-
import models400
from models400 import *
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.cell.cell import  Cell

import random
from pymysql.err   import  IntegrityError
from sqlalchemy.exc import IntegrityError
from sqlalchemy import insert

"""清除脚本
TBD
说明：
4个tab红色表头是必填
m_area 不存在的才写入
m_site 不存在的才写入
m_room 不存在的才写入，机房已经存在且机房内已经有设备return
m_device 每个机房写表格中配置的所有设备类型
m_signal 按照设备类型和m_device的首列匹配设备测点进行写入
"""
# excel_file = './lsc400_data_dcim.xlsx'
excel_file = './lsc400_data_gx_.xlsx'
#excel_file = './lsc400_data_guangdong.xlsx'
#excel_file = './lsc400_data_guangdong.xlsx'
#excel_file = './lsc400_data_guangdong.xlsx'
#excel_file = './lsc400_data_guangdong.xlsx'
roomtype_list=[1,2,11,12,13,14,51,52,53,54,55]
sitetype_list=[1,2,3,4]# DATACENTER =1	数据中心 ROOM＝2	通信机楼 LOCALTRANS=3	传输节点  STATION=4	通信基站
d_device_id =43031 # 40002 SELECT   *   FROM    m_device   WHERE DeviceID>42405
signals=[]
def load_xls_area():
    wb = load_workbook(excel_file)
    ws=wb['m_area']
    result_list=[]
    for row in ws.iter_rows(min_row=2):
        scid = row[0].value
        areaid = row[1].value
        lastareaid = row[2].value
        areaname = row[3].value
        #print('%s %s %s %s %s %s %s %s' % (fsuid, fsuver,siteid, sitename,roomid,roomname,interval, m))
        if scid != 'None':
            result_list.append([scid,areaid,lastareaid,areaname])
    return result_list

def load_xls_room():
    wb = load_workbook(excel_file)
    ws=wb['m_room']
    result_list=[]
    for row in ws.iter_rows(min_row=2):
        scid = row[0].value
        roomID = row[1].value
        siteID = row[2].value
        roomType = row[3].value
        roomName = row[4].value
        roomDesc = row[5].value

        #print('%s %s %s %s %s %s %s %s' % (fsuid, fsuver,siteid, sitename,roomid,roomname,interval, m))
        if scid != 'None' and roomID is not None and siteID is not None:
            result_list.append([scid,roomID,siteID,roomType,roomName,roomDesc])
    return result_list

def load_xls_site():
    global excel_file
    wb = load_workbook(excel_file)
    ws=wb['m_site']
    result_list=[]
    for row in ws.iter_rows(min_row=2):
        SCID = row[0].value
        SiteID = row[1].value
        SiteName = row[2].value
        SiteDesc = row[3].value
        Longitude = row[4].value
        latitude = row[5].value
        NodeFeatures = row[6].value
        SiteType = row[7].value
        AreaId = row[8].value
        if SiteID != 'None':
            result_list.append([SCID,SiteID,SiteName,SiteDesc,Longitude,latitude,NodeFeatures,SiteType,AreaId])
    return result_list

def load_xls_device():
    global excel_file
    wb = load_workbook(excel_file)
    ws=wb['m_device']
    result_list=[]
    for row in ws.iter_rows(min_row=2):
        m = row[0].value
        SCID = row[1].value
        DeviceID = row[2].value
        RoomID = row[3].value
        SiteID = row[4].value
        DeviceName = row[5].value
        DeviceDesc = row[6].value
        DeviceType = row[7].value
        Productor = row[8].value
        Version = row[9].value
        DeviceModel = row[10].value
        LocateNeStatus = row[11].value
        ModelId = row[12].value


        if SiteID != 'None':
            result_list.append([m,SCID,DeviceID,RoomID,SiteID,DeviceName,DeviceDesc,DeviceType,Productor,Version,DeviceModel,LocateNeStatus,ModelId])
    return result_list

def load_xls_signal():
    global excel_file
    wb = load_workbook(excel_file)
    ws=wb['m_signal']
    result_list = []
    for row in ws.iter_rows(min_row=2):
        m =row[0].value
        signal_SCID = row[1].value
        signal_SiteID = row[2].value
        signal_DeviceID = row[3].value
        #  ALARM=0	告警   其他是 1、2、3、4  数字输出量，遥控 | 2 模拟输出量，遥调| 3  模拟输入量，遥测| 4字输入量（包含多态数字输入量）遥信
        signal_Type = row[4].value
        signal_SignalID = row[5].value
        signal_SignalNumber = row[6].value
        signal_SignalName = row[7].value
        signal_AlarmLevel = row[8].value
        signal_Threshold = row[9].value
        signal_StoragePeriod = row[10].value
        signal_AbsoluteVal = row[11].value
        signal_RelativeVal = row[12].value
        signal_StaticVal = row[13].value
        signal_Describe = row[14].value
        signal_NMAlarmID = expr_none(row[15].value)
        if signal_AlarmLevel is None:
            signal_AlarmLevel = '4'

        result_list.append([m,signal_SCID,signal_SiteID,signal_DeviceID,signal_Type,signal_SignalID ,
                                     signal_SignalNumber,signal_SignalName,signal_AlarmLevel ,signal_Threshold ,
                                     signal_StoragePeriod,signal_AbsoluteVal ,signal_RelativeVal ,signal_StaticVal ,
                                     signal_Describe ,signal_NMAlarmID])
    return result_list

def expr_none(somestr):
    if somestr == None:
        return ''
    else:
        return somestr

def add_area(area_list):
    m_area = MArea()
    m_area.SCID = area_list[0]
    m_area.AreaID = area_list[1]
    m_area.LastAreaID = area_list[2]
    m_area.AreaName =  area_list[3]
    session.add(m_area)

def add_room(room_list):
    m_room = MRoom()
    m_room.SCID = room_list[0]
    m_room.RoomID = room_list[1]
    m_room.SiteID = room_list[2]
    m_room.RoomType = room_list[3]
    m_room.RoomName = room_list[4]
    m_room.RoomDesc = room_list[4]+"desc"
    session.add(m_room)
    session.commit()

def add_site(site_list):
    m_site = MSite()
    m_site.SCID = site_list[0]
    m_site.SiteID = site_list[1]
    m_site.SiteName = site_list[2]
    m_site.SiteDesc = str(site_list[2])+"SiteDesc"
    m_site.Longitude = site_list[4]
    m_site.Latitude = site_list[5]
    m_site.NodeFeatures = site_list[6]
    m_site.SiteType = site_list[7]
    m_site.AreaId = site_list[8]
    session.add(m_site)
    try:
        session.commit()
    except IntegrityError as e:
        print("Error: Division by zero is not allowed")

    finally:
        print("The 'finalil' block was executed successfully --add_site")

def add_device(device_list):
    m_device = MDevice()
    m_device.SCID = device_list[0]
    m_device.DeviceID = device_list[1]
    m_device.RoomID = device_list[2]
    m_device.SiteID = device_list[3]
    m_device.DeviceName = device_list[4]
    m_device.DeviceDesc = device_list[5]
    m_device.DeviceType = device_list[6]
    m_device.Productor = device_list[7]
    m_device.Version = device_list[8]
    m_device.DeviceModel = device_list[9]
    m_device.LocateNeStatus = device_list[10]
    m_device.ModelId = device_list[11]
    session.add(m_device)

def add_m_signal(signals):
    ins = insert(t_m_signal).values(signals)

    with engine.connect() as connection:
        cursor = connection.execute(ins)
    cursor.close()
    connection.close()

def setDeviceMax():
    global  d_device_id
   #获取到最大的设备ID global d_device_id
    if  (queryDeviceMax())[0][0] is not None :
        d_device_id=queryDeviceMax()[0][0]+1

def test():
    area_list = load_xls_area()
    # logger.debug('["area_list区域列表 %s" \n\t',area_list)
    areaID_filter_list = [area[1] for area in area_list]
    # logger.debug('["areaID_filter_listt表格中区域ID列表 %s" \n\t', areaID_filter_list)
    site_list = load_xls_site()
    # logger.debug('["site_list %s" \n\t', site_list)
    siteID_filter_list = [site[1] for site in site_list if (site[-1] in areaID_filter_list)]
    # logger.debug('["siteID_filter_list %s" \n\t', siteID_filter_list)
    room_list = load_xls_room()
    # logger.debug('["room_list" %s \n\t', room_list)
    roomID_filter_list = [room for room in room_list if (room[2] in siteID_filter_list)]
    # logger.debug('["roomID_filter_list表格中机房ID列表" %s \n\t', roomID_filter_list)
    device_list = load_xls_device()
    # logger.debug('["device_list %s" \n\t', device_list)
    signal_list = load_xls_signal()
    # logger.debug('["signal_list %s" \n\t', signal_list)
    for area in area_list:
        sCID = area[0]
        areaID = area[1]
        lastAreaID = area[2]
        areaName = area[3]
        print(sCID,areaID)
        #如果区域已经存在就不写入了
        if  models400.queryMArea(sCID,areaID) is None :
            add_area([area[0], area[1], area[2], area[3]])
            logger.debug('["区域写入 " \n\t')
        else:
            logger.debug('["区域已经存在sCID, areaID %s %s" \n\t', sCID, areaID)
        station_filter_list = [station for station in site_list if (station[-1] == areaID)]
        print("station_filter_list================================:%s" % station_filter_list)
        if station_filter_list==[] or (len(station_filter_list)!=len(site_list)) :
             logger.debug('["m_site-ID与m_area不全对应的 %s %s" \n\t', sCID, areaID)
             return
        for station in station_filter_list:
            station_id = station[1]
            print([station[0], station[1],station[2],str(station[2])+"desc",station[4],station[5],station[6],station[7],station[8] ])
            #如果站点已经存在不写入
            if models400.queryMSitefist(station[0], station[1]) is None:
                add_site([station[0], station[1],station[2],str(station[2])+"desc",station[4],station[5],station[6],station[7],station[8] ])
                logger.debug('["站点写入" \n\t')
            else:
                logger.debug('["站点已经存在SCID	SiteID %s %s" \n\t', station[0], station[1])
        #  m_room的站点ID在m_siteID在sheet存在
        room_filter_list = [room for room in room_list if (room[2] in siteID_filter_list)]
        print("room_filter_list====================",room_filter_list)
        if  len(room_filter_list)==0  or (len(room_filter_list)!=len(room_list)):
            logger.debug('["m_room的站点ID在m_sitesheet不全存在,请修改表" \n\t')
            return
        #开始写入机房列表，不存在的会写到
        for room in room_filter_list:
            print([room[0], room[1], room[2], room[3], room[4], room[5] ])
            # 如果站点已经存在不写入
            if models400.queryMRoomfist(room[0], room[1], room[2]) is None:
                add_room([room[0], room[1], room[2], room[3], room[4] ])
                logger.debug('["机房不存在，进行写入 " \n\t')
            else:
                # pass
                logger.debug('["机房已经存在SCID RoomID SiteID %s %s %s" \n\t', room[0], room[1], room[2])
                if( models400.queryDevicecount(room[1]))[0][0]==0:
                    logger.debug("机房已经存在但是机房内还没有设备-继续写入设备")
                    pass
                    #如果机房存在但是还没有设备才继续写入
                else:
                    logger.debug("机房已经存在且已有设备，为避免重复写入，请确定机房内无设备")
                    return
        #向每个机房写如相同的设备数据
        device_filter_list = [device for device in device_list ]
        # device_filter_list = [device for device in device_list if device[1] in  roomID_filter_list]
        if len(device_filter_list) == 0 :
            logger.debug('["m_device表格为空" \n\t')
            return
        #每个接卸写入数据前获取一次中间库最 deviceID setDeviceMax()
        setDeviceMax()
        for room in room_filter_list:
            SCID=room[0]
            RoomID=room[1]
            SiteID=room[2]
            global d_device_id
            for device in device_filter_list:
                m_device = device[0]
                dSCID = SCID
                dDeviceID = d_device_id
                d_device_id=d_device_id+1
                dRoomID = RoomID
                dSiteID =SiteID
                if (device[5] is not None):
                    dDeviceName = device[5]
                dDeviceDesc = str(device[5])+"Desc"
                if device[7] is not None :
                    dDeviceType = device[7]
                if device[8]   is not None :
                    dProductor = device[8]
                #表格未填就给null
                if  device[9] is None:
                    dVersion =   None
                if device[10] is None:
                 dDeviceModel = None
                if device[11] is None:
                    dLocateNeStatus = None
                if device[12] is None:
                    dModelId = None
                print("开始写device" )
                add_device([dSCID,dDeviceID,dRoomID,dSiteID,dDeviceName,dDeviceDesc,dDeviceType,dProductor,dVersion,dDeviceModel,dLocateNeStatus,dModelId])
                signal_filter_list = [signal for signal in signal_list]
                signal_m = signal_list[0][0]
                # 后续可定制某设备特定测点组合
                if m_device == signal_m:
                    logger.debug('["开始写入测点编码m_device,signal_m %s %s" \n\t', m_device, signal_m)
                    global signals
                    signals=[]
                    pass
                    for signal in signal_filter_list :
                        signal_SCID = dSCID
                        signal_SiteID = dSiteID
                        signal_DeviceID = dDeviceID
                        #  ALARM=0	告警   其他是 1、2、3、4  数字输出量，遥控 | 2 模拟输出量，遥调| 3  模拟输入量，遥测| 4字输入量（包含多态数字输入量）遥信
                        signal_Type = signal[4]
                        signal_SignalID = signal[5]
                        signal_SignalNumber = signal[6]
                        signal_SignalName = signal[7]
                        signal_AlarmLevel = signal[8]
                        signal_Threshold = signal[9]
                        signal_StoragePeriod = signal[10]
                        signal_AbsoluteVal = signal[11]
                        signal_RelativeVal = signal[12]
                        signal_StaticVal = signal[13]
                        signal_Describe = signal[14]
                        signal_NMAlarmID = signal[15]
                        signaone={"SCID" : signal_SCID,
                            "SiteID" : signal_SiteID,
                            "DeviceID" :signal_DeviceID,
                            "Type" :  signal_Type,
                            "SignalID" :  signal_SignalID,
                            "SignalNumber" :  signal_SignalNumber,
                            "SignalName" :  signal_SignalName,
                            "AlarmLevel" :  signal_AlarmLevel,
                            "Threshold" :  signal_Threshold,
                            "StoragePeriod" : signal_StoragePeriod,
                            "AbsoluteVal" :  signal_AbsoluteVal,
                            "RelativeVal" : signal_RelativeVal,
                            "StaticVal" :  signal_StaticVal,
                            "Describe" :  signal_Describe,
                            "NMAlarmID" : signal_NMAlarmID}
                        if  str(dDeviceType).zfill(3) == signal_SignalID[0:3]:
                            #找到设备类型对应的测点编码开始组装
                            signals.append(signaone)
                    logger.debug("开始写入signal")
                    logger.debug('["一个设备的测点集合写一次 %s  " \n\t', signals)
                    add_m_signal(signals)
        session.commit()


if __name__=="__main__":
    #前置站点准备 新增站点-增加映射
    test()

