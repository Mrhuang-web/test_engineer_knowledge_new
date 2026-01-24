# encoding:utf-8
"""
@CreateTime:      2023/9/25 11:11
@Author:          Tsuiguangchun
@FileName:        all_excelCase_file_path.py
@IDE_SoftWare:    PyCharm
@description:     全部测试用例数据文件路径，将文件路径统一放到一个地方，方便日后维护
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
"""
import json
import os
import time

import asptest.service.datadriver as datadriver
import openpyxl

ROOT_PATH = os.path.dirname(os.path.dirname(__file__))


def load_excel_data(dir_name, file_name, file_suffix=".xlsx", index=0, case_mark='', mark_column=-2, dir_name2=None):
    """
    dir_name ,dir_name2 文件目录/子目录名称
    file_name 文件名称
    index: 读取excel表页索引，即读取第几页的数据
    case_mark: 用例标记，若传入值，则只有包含此标记的用例才会执行，其余用例数据直接过滤，不执行
    mark_column: 用例标记对应的列序号，默认为倒数第2列，可以自行指定
    """
    if dir_name2:
        file_path = os.path.join(ROOT_PATH, dir_name, dir_name2, file_name + file_suffix)
        print(f"文件路径：{file_path}")
    else:
        file_path = os.path.join(ROOT_PATH, dir_name, file_name + file_suffix)
    data = datadriver.load_data_from_excel(file_path, index, case_mark, mark_column)
    # print(data)
    return data


class ExcelWriter:
    def __init__(self, dir_name, file_name, file_suffix=".xlsx",
                 dir_name2=None, headers=None, new_workbook=False):
        if dir_name2:
            self.filename = os.path.join(ROOT_PATH, dir_name, dir_name2, file_name + file_suffix)
            print(f"文件路径：{self.filename}")
        else:
            self.filename = os.path.join(ROOT_PATH, dir_name, file_name + file_suffix)
        print(f"文件路径：{self.filename}")
        if new_workbook:
            self.workbook = openpyxl.Workbook()
        self.workbook = openpyxl.load_workbook(self.filename)
        self.sheet = self.workbook.active
        self.sheet_num = 0
        self.headers = headers
        self.current_row = 2

    def add_sheet(self, name, headers=None):
        if headers is None:
            headers = self.headers
        else:
            self.headers = headers

        self.sheet = self.workbook.create_sheet(name, self.sheet_num)
        if not (self.headers is None and headers is None):
            for i, t in enumerate(headers):
                self.sheet.cell(1, i + 1, t)
            self.current_row += 1
            self.sheet_num += 1

    def remove_sheet(self, sheet_name):
        """
        删除指定表
        """
        self.workbook.remove(sheet_name)

    def save(self):
        self.workbook.save(self.filename)

    def write(self, data, new_sheet=False):
        if not hasattr(self, 'headers'):
            raise ValueError('please set headers before write data into excel.')
        if not hasattr(self, 'sheet'):
            self.add_sheet('Sheet1')

        if not self.headers is None and len(data) != len(self.headers):
            raise ValueError('长度不匹配')
        if isinstance(data, dict):
            for i, t in enumerate(self.headers):
                if data[t] == None:

                    self.sheet.cell(self.current_row, i + 1, '')

                else:
                    self.sheet.cell(self.current_row, i + 1, data[t])
        elif isinstance(data, tuple) or isinstance(data, list):
            for i, t in enumerate(data):
                if t == None:
                    self.sheet.cell(self.current_row, i + 1, '')
                else:
                    self.sheet.cell(self.current_row, i + 1, t)
        self.current_row += 1


def write_to_excel(dir_name, file_name, value, file_suffix=".xlsx", index="Sheet1", exits=True, row=None, column=None,
                   dir_name2=None):
    """
    dir_name ,dir_name2 文件目录/子目录名称
    file_name 文件名称
    index: 读取excel表页索引，即读取第几页的数据
    row ;: 行
    column ：： 列
    value ：： 写入的内容，列表或元组，需要嵌套列表：[[1,2],[3,4]]，指定单个单元格写入字符串:"1"
    exits :: 默认True ,如果传Flase 新建
    """
    # 打开现有的Excel文件
    if dir_name2:
        file_path = os.path.join(ROOT_PATH, dir_name, dir_name2, file_name + file_suffix)
        print(f"文件路径：{file_path}")
    else:
        file_path = os.path.join(ROOT_PATH, dir_name, file_name + file_suffix)
    if not exits:
        work_book = openpyxl.Workbook()

    else:
        workbook = openpyxl.load_workbook(file_path)
        # 获取活动工作表
        worksheet = workbook.active
        # 获取指定名称的工作表
        # worksheet = workbook[index]
        if isinstance(value, list):
            # 循环遍历每一行数据，并逐个写入单元格
            value = tuple(value)
            for row in value:

                if isinstance(row, list):
                    # row = tuple(row)
                    worksheet.append(row)
                    workbook.save(file_path)

        # 指定要写入的单元格范围，并逐个写入数据
        # worksheet.cell(row=row, column=column, value=value)  # 第一列，第i行，值row[0]
        # time.sleep(1)
        # workbook.save(file_path)
    #
    #         worksheet.cell(row=row, column=2, value=row[1])  # 第二列，第i行，值row[1]
    #         worksheet.cell(row=row, column=3, value=row[2])  # 第三列，第i行，值row[2]
    #         i += 1  # i自增1，用于迭代不同行数
    # return data


def read_sql(dir_name, dir_name2=None, file_name=None, file_suffix=".sql"):
    """
    同一个sql文件,多条sql,用；进行分割
    """
    file_path = os.path.join(ROOT_PATH, dir_name, dir_name2, file_name + file_suffix)
    with open(file_path, 'r') as file:
        sql_script = file.read().split(';')
        if len(sql_script) >= 2:
            print(sql_script)
            return sql_script
        else:
            print(sql_script)
            return sql_script[0]

        # print(f"sql_script---->{sql_script},{len(sql_script)}")
        # return sql_script


def read_json_file(dir_name, file_name=None, file_suffix=".json", **kwargs):
    if kwargs:
        file_path = os.path.join(ROOT_PATH, dir_name, str(kwargs), file_name + file_suffix)
    else:
        file_path = os.path.join(ROOT_PATH, dir_name, file_name + file_suffix)
    with open(file_path, 'r', encoding='utf-8') as f:
        try:
            json_data = json.load(f)
            print(json_data)
            print(type(json_data))
            return json_data
        except json.JSONDecodeError as e:
            print(f"Error decoding JSON file: {e}")
            return None


# def read_json_file(dir_name, file_name=None, file_suffix=".json",**kwargs):
#     if kwargs:
#         file_path = os.path.join(ROOT_PATH, dir_name, str(kwargs), file_name + file_suffix)
#     file_path = os.path.join(ROOT_PATH, dir_name, file_name + file_suffix)
#     with open(file_path, 'r') as f:
#         json_data = json.load(f)
#         print(json_data)
#         print(type(json_data))
#         return json_data

"""
~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~             
"""

