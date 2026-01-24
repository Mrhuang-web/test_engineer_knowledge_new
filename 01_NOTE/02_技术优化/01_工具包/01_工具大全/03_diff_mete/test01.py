import openpyxl


def compare_columns(excel_file):
    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Sheet1']

    # 获取C列和I列的数据（从第二行开始，跳过标题行）
    c_column = []
    i_column = []

    for row in range(2, sheet.max_row + 1):
        c_value = sheet.cell(row=row, column=3).value  # C列
        i_value = sheet.cell(row=row, column=8).value  # I列

        if c_value is not None:
            # 确保C列的值是字符串格式，保留前导零
            c_column.append(str(c_value).zfill(6) if isinstance(c_value, (int, float)) else str(c_value))
        if i_value is not None:
            # 确保I列的值是字符串格式，保留前导零
            i_column.append(str(i_value).zfill(6) if isinstance(i_value, (int, float)) else str(i_value))

    # 找出C列有但I列没有的值
    c_only = set(c_column) - set(i_column)

    # 找出I列有但C列没有的值
    i_only = set(i_column) - set(c_column)

    # 将结果写入L列和M列
    l_row = 2
    m_row = 2

    for row in range(2, sheet.max_row + 1):
        c_value = sheet.cell(row=row, column=3).value
        if c_value is not None:
            # 确保比较时格式一致
            c_str = str(c_value).zfill(6) if isinstance(c_value, (int, float)) else str(c_value)
            if c_str in c_only:
                sheet.cell(row=l_row, column=12).value = c_str  # L列
                l_row += 1

        i_value = sheet.cell(row=row, column=9).value
        if i_value is not None:
            # 确保比较时格式一致
            i_str = str(i_value).zfill(6) if isinstance(i_value, (int, float)) else str(i_value)
            if i_str in i_only:
                sheet.cell(row=m_row, column=13).value = i_str  # M列
                m_row += 1

    # 保存文件
    wb.save('./result/comparison_result.xlsx')
    print("比较完成，结果已保存到 comparison_result.xlsx")


# 使用示例
compare_columns('excel/对比文件1.xlsx')