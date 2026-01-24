#-*- coding:utf-8 -*-
from module import *
from openpyxl import Workbook
from openpyxl import load_workbook

fsuid = '202407091142'

"""清除脚本
SET @fsuid = '100000000000009';
DELETE FROM signals WHERE fsuid=@fsuid;
DELETE FROM device WHERE fsuid=@fsuid;
DELETE FROM fsu WHERE fsuid=@fsuid;
"""

def load_xls_fsu():
    wb = load_workbook('./fsu_data.xlsx')
    ws=wb['fsu']
    result_list=[]
    for row in ws.iter_rows(min_row=2):
        fsuid = row[0].value
        fsuname = row[1].value
        fsuver = row[2].value
        siteid = row[3].value
        sitename = row[4].value
        roomid = row[5].value
        roomname = row[6].value
        interval = row[7].value
        m = row[8].value
        #print('%s %s %s %s %s %s %s %s' % (fsuid, fsuver,siteid, sitename,roomid,roomname,interval, m))
        if fsuid != 'None' and m!='None':
            result_list.append([fsuid, fsuname, fsuver,siteid, sitename,roomid,roomname,interval, m])
    return result_list
    
def load_xls_device():
    wb = load_workbook('./fsu_data.xlsx')
    ws=wb['device']
    result_list = []
    for row in ws.iter_rows(min_row=2):
        m = row[0].value
        fsuid = row[1].value
        deviceid = row[2].value
        devicename = row[3].value
        devdescribe = row[4].value
        siteid = row[5].value
        sitename = row[6].value
        roomid = row[7].value
        roomname = row[8].value
        devicetype = row[9].value
        devicesubtype = row[10].value
        model =row[11].value
        brand = row[12].value
        ratedcapacity = row[13].value
        version = row[14].value
        beginruntime = expr_none(row[15].value)
        confremark = expr_none(row[16].value)
        #print('%s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s' % (m,fsuid, deviceid,devicename, devdescribe,siteid,sitename,roomid,roomname,devicetype,devicesubtype,model,brand,ratedcapacity,version,beginruntime,confremark))
        result_list.append([m,fsuid, deviceid,devicename, devdescribe,siteid,sitename,roomid,roomname,devicetype,devicesubtype,model,brand,ratedcapacity,version,beginruntime,confremark])
    return result_list
    
def load_xls_signal():
    wb = load_workbook('./fsu_data.xlsx')
    ws=wb['signal']
    result_list = []
    for row in ws.iter_rows(min_row=2):
        m = row[0].value
        deviceid = row[1].value
        Type = row[2].value
        ID = row[3].value
        SignalName = row[4].value
        SignalNumber = row[5].value
        AlarmLevel = row[6].value
        Threshold = row[7].value
        NMAlarmID = expr_none(row[8].value)
        #print('%s %s %s %s %s %s %s %s %s' % (m, deviceid,Type,ID,SignalName,SignalNumber,AlarmLevel,Threshold,NMAlarmID))
        result_list.append([m, deviceid,Type,ID,SignalName,SignalNumber,AlarmLevel,Threshold,NMAlarmID])
    return result_list

def expr_none(somestr):
    if somestr == None:
        return ''
    else:
        return somestr

def add_fsu(fsu_xls_record):
    fsu = Fsu()
    fsu.fsuid = fsu_xls_record[0]
    fsu.fsuname = fsu_xls_record[1]
    fsu.username = fsu_xls_record[0]
    fsu.password = fsu_xls_record[0]
    fsu.fsuip = '10.1.24.7'
    fsu.fsumac = '01-01-01-01'
    fsu.fsuver = fsu_xls_record[2]
    fsu.siteid = fsu_xls_record[3]
    fsu.sitename = fsu_xls_record[4]
    fsu.roomid= fsu_xls_record[5]
    fsu.roomname = fsu_xls_record[6]
    fsu.tfsustatus_cpuusage = '10'
    fsu.tfsustatus_memusage = '20'
    fsu.tfsustatus_harddiskusage = '30'
    fsu.interval = fsu_xls_record[7]
    session.add( fsu )
    session.commit()

def add_device(fsuid,fsuname,device_xls_record):
    device = Device()
    device.fsuid = fsuid        #??fsuid, ??excel??????excel????
    device.deviceid = device_xls_record[2] 
    device.devicename = device_xls_record[3]
    device.siteid = device_xls_record[5]
    device.roomid =  device_xls_record[7]
    device.sitename =  device_xls_record[6]
    device.roomname =  device_xls_record[8]
    device.devicetype =  device_xls_record[9]
    if device.devicetype == 76:
        device.devicename = fsuname    # 对于type=76的，设备名称改为fsu的名称
    device.devicesubtype =  device_xls_record[10]
    device.model =  device_xls_record[11]
    device.brand =  device_xls_record[12]
    device.ratedcapacity =  device_xls_record[13]
    device.version =  device_xls_record[14]
    device.beginruntime =  device_xls_record[15]
    device.devdescribe =  device_xls_record[4]
    device.confremark =  device_xls_record[16]
    session.add( device )
    #session.commit()

def add_signal(fsuid,deviceid,signal_xls_record):
    signal = Signals()
    signal.fsuid = fsuid
    signal.deviceid = deviceid
    signal.signalsid = signal_xls_record[3].zfill(6)    #???6??,???mete_code
    signal.type = signal_xls_record[2] 
    signal.signalname = signal_xls_record[4]
    #print(signal_xls_record[5])
    signal.signalnumber =  str(signal_xls_record[5]).zfill(3)  #???3?????000
    signal.alarmlevel =  signal_xls_record[6]
    signal.thresbhold =  signal_xls_record[7]
    signal.nmalarmid =  signal_xls_record[8]
    signal.measuredval =  ''
    signal.setupval =  ''
    signal.status =  ''
    signal.time =  '2019-06-25 00:00:00'
    session.add( signal )
    #session.commit()
    

def test():
    #global fsuid
    #print queryFsu(fsuid)
    #if queryFsu(fsuid) == None:
    #    add_fsu()
    fsu_list = load_xls_fsu()
    device_list = load_xls_device()
    signal_list = load_xls_signal()
    
    for fsu_record in fsu_list:
        fsuid = fsu_record[0]
        fsuname = fsu_record[1]
        print(queryFsu(fsuid))
        if queryFsu(fsuid) == None:
            m_idx = fsu_record[8]
            add_fsu(fsu_record)
            print(fsu_record)
            dev_list = [li for li in device_list if li[0] == m_idx]
            for device_record in dev_list:
                devid = device_record[2]
                add_device(fsuid,fsuname,device_record)
                print(device_record)
                sig_list = [li for li in signal_list if li[0] == m_idx and li[1] == devid]
                for signal_record in sig_list:
                    sigid = signal_record[1]
                    add_signal(fsuid, devid, signal_record)
                    print(signal_record)
    session.commit()
    
if __name__=="__main__":
    test()

