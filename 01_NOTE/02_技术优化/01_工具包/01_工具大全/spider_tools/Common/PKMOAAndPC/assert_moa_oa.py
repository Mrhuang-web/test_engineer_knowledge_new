# -*- coding: utf-8 -*-
"""
@Time ： 2024/12/16 14:09
@Auth ： sunzhonghua
@File ：tt.py.py
@IDE ：PyCharm
"""
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
import requests
import json
import jmespath
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


moa_headers = {
    'Host': 'fee.net.chinamobile.com' ,
    'Connection': 'keep-alive' ,
    'Pragma': 'no-cache' ,
    'Cache-Control': 'no-cache' ,
    'sec-ch-ua': '"Chromium";v="116", "Not)A;Brand";v="24", "Android WebView";v="116"' ,
    'sec-ch-ua-mobile': '?1' ,
    'authorization': 'cc94f04a-a1a1-49c1-8694-53230a8fbbf1-a',
    'User-Agent': 'Mozilla/5.0 (Linux; Android 13; V2239A Build/TP1A.220624.014; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/116.0.0.0 Mobile Safari/537.36' ,
    'Content-Type': 'application/json' ,
    'Accept': 'application/json, text/plain, */*' ,
    'head_orgAccount': 'alauda' ,
    'head_userName': 'sunzhonghua' ,
    'sec-ch-ua-platform': '"Android"' ,
    'Origin': 'https://fee.net.chinamobile.com' ,
    'X-Requested-With': 'com.cmiot.eoms' ,
    'Sec-Fetch-Site': 'same-origin' ,
    'Sec-Fetch-Mode': 'cors' ,
    'Sec-Fetch-Dest': 'empty' ,
    'Referer': 'https//fee.net.chinamobile.com/spider/moa/h5/' ,
    'Accept-Encoding': 'gzip, deflate, br' ,
    'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7' ,
    'Cookie': 'JSESSIONID=E5EC3631E574489FE1967D854F32B16555; portal_token=cc94f04a-a1a1-49c1-8694-53230a8fbbf1-a'
}
pcoa_headers = {
    'Connection': 'keep-alive' ,
    'Authorization': '59PWo5VanZqS8XcQEBBoiH5PR7I3iEYAENK25YXjxGZeJbmx77mGoTdHo5fMToMN' ,
    'Content-Type': 'application/json' ,
    'Accept': 'application/json, text/plain, */*' ,
    'head_orgAccount': 'alaudaaaaa' ,
    'head_userName': 'sunzhonghuaaaaa' ,
    }

def send(url="",methods="POST", requestdata="",jsonpath="",headers={}):
    """
    url="" moa或者OA的url
    methods="POST" "GET"
    PCrequest="",
    jsonpath=""
    """
    # url = url
    # print(url)
    # print(methods)
    # print(requestdata)
    # print(jsonpath)
    data=[]
    if methods =="POST":
        data = json.loads(requestdata)
        response = requests.post(url, headers=headers, json=data, verify=False)  # verify=False 相当于 --insecure
    if methods =="GET":
        response = requests.get(url, headers=headers,json=data,verify=False)  # verify=False 相当于 --insecure
    if response.status_code == 200:
        data = response.json()
        if  str(jsonpath).isdigit():
            "{}".format(str(jsonpath))
            getdata = data[ "{}".format(str(jsonpath))]
        else:
            getdata =jmespath.search(jsonpath, data)
        return getdata
    else:
        print("返回数据失败响应码为response.status_code" + str(response.status_code))
def left_pk_right(leftdata,rightdata):
    # 检查否都不为空
    if leftdata and rightdata:
        if str(leftdata) == str(rightdata):
            return "pass"
        else:
            return str(leftdata) + "|" + str(rightdata)
    else:
        return "未获取到数据"

def ispass():
    # 加载Excel文件
    workbook = load_workbook(filename='moa_data_assert.xlsx')
    # 选择活动的工作表
    sheet = workbook.active
    # 定义绿色填充和字体
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    font = Font(color='000000')  # 黑色字体
    # 遍历工作表中的所有行
    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=False):
        # print(row[1].value)
        # print(row[2].value)
        # print(row[3].value)
        # 检查第8列（即索引7）的单元格是否存在
        if len(row) > 11:
            leftd = send(url=row[1].value , methods=row[2].value, requestdata=row[3].value, jsonpath=row[5].value,headers=moa_headers)
            rightd = send(url=row[7].value , methods=row[8].value, requestdata=row[9].value, jsonpath=row[11].value,headers=pcoa_headers)
            print(leftd, rightd)
            print(type(leftd), type(rightd))
            ispassv = left_pk_right(leftd, rightd)
            if ispassv =="pass":
                row[12].value = ispassv
                row[12].fill = green_fill
                row[12].font = font
            else:
                row[12].value = ispassv
                row[12].fill = red_fill
                row[12].font = font

    # 保存修改后的工作簿
    workbook.save(filename='moa_data_assert.xlsx')
    # 关闭工作簿
    workbook.close()

if __name__ == '__main__':
    ispass()