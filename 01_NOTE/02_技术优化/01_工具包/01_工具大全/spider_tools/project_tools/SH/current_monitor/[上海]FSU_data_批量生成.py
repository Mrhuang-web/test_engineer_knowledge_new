# 使用前请安装: pip install openpyxl
from openpyxl import Workbook
import os
import time
import requests
import spider_tools.Common.TestEnv as TestEnv
from asptest.common.mysqlpool import MySQLHelper

ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
work = os.path.join(ROOT_PATH, "Params", "upload", "room_name.csv")
conn_obj = MySQLHelper(TestEnv.ServerConfig['db_pas_sh'])

# 配置参数
CONFIG = {
    'fsu_ver': 'V1.0',
    'interval': 60,
    'm': '通用',
    'brand': '通用',
    'version': '1.0',
    'devicesubtype': 1,
    'ratedcapacity': 0.000000,
    'signal_channels': 5,  # 每个测点生成的通道号数量
}


# ... device_tuple 函数保持不变 ...
def device_tuple():
    """设备类型定义和测点定义"""
    device = {
        1: '高压配电柜', 2: '低压交流配电', 3: '变压器', 4: '低压直流配电',
        5: '发电机组', 6: '开关电源', 7: '铅酸电池', 8: 'UPS设备',
        9: 'UPS配电', 11: '机房专用空调', 12: '中央空调末端', 13: '中央空调主机',
        14: '变换设备', 15: '普通空调', 16: '极早期烟感', 18: '电池恒温箱', 68: '锂电池',
        76: '动环监控', 77: '智能通风换热', 78: '风光设备', 87: '高压直流电源', 88: '高压直流电源配电',
        92: '智能电表', 17: '机房环境', 93: '智能门禁',
    }

    device_mete = {
        1: [['001101', '2', '模块01开机'], ['001102', '2', '模块01关机'], ['001307', '1', '相电流Ia'],
            ['001304', '1', '相电压Ua'], ['001311', '1', '有功功率Pa']],
        2: [['002101', '2', '低压开关合闸'], ['002102', '2', '低压开关分闸'], ['002307', '1', '有功功率Pa'],
            ['002304', '1', 'A相功率因数'], ['002311', '1', 'B相功率因数']],
        3: [['003101', '2', '风机关控制'], ['003102', '2', '风机开控制'], ['003301', '1', 'A相温度'],
            ['003302', '1', 'B相温度'], ['003303', '1', 'C相温度']],
        4: [['004101', '2', '直流电压'], ['004102', '2', '直流电流'], ['004305', '1', '输入01电能'],
            ['004306', '1', '分路01电能'], ['004307', '1', '分路01功率']],
        5: [['005101', '2', '高压断路器合闸'], ['005102', '2', '高压断路器分闸'], ['005307', '1', '输出线电压Uab'],
            ['005308', '1', '输出线电压Ubc'], ['005309', '1', '输出线电压Uca']],
        6: [['006101', '2', '模块01开机'], ['006102', '2', '模块01关机'], ['006301', '1', '输入相电压Ua'],
            ['006302', '1', '输入相电压Ub'], ['006303', '1', '输入相电压Uc']],
        7: [['007101', '1', '标示单体01温度'], ['007301', '1', '单体01电压'], ['007302', '1', '单体01温度'],
            ['007303', '1', '单体01内阻']],
        8: [['008101', '2', '紧急停机'], ['008301', '1', '输入线电压Uab'], ['008302', '1', '输入线电压Ubc'],
            ['008303', '1', '输入线电压Uca']],
        9: [['009301', '1', '分路01相电流Ia'], ['009302', '1', '分路01相电流Ib'], ['009303', '1', '分路01相电流Ic']],
        11: [['011101', '2', '模块01开机'], ['011102', '2', '模块01关机'], ['011301', '1', '回风温度'],
             ['011302', '1', '回风湿度'], ['011303', '1', '工作电流']],
        12: [['012301', '1', '回风温度'], ['012302', '1', '回风湿度'], ['012303', '1', '工作电流']],
        13: [['013301', '1', '线电压Uab'], ['013302', '1', '线电压Ubc'], ['013303', '1', '线电压Uca']],
        14: [['014301', '1', '输入电压'], ['014302', '1', '输出电压'], ['014303', '1', '主机温度']],
        15: [['015101', '2', '开机'], ['015102', '2', '关机'], ['015201', '1', '温度设定'], ['015202', '1', '风量设定'],
             ['015203', '1', '回风温度设定']],
        16: [['016301', '1', '01管烟雾浓度'], ['016302', '1', '烟感01气流回流率'], ['016303', '1', '报警灵敏度']],
        18: [['018101', '2', '蜂鸣器开关'], ['018102', '2', '存储记录开关'], ['018301', '1', '设备系统时间'],
             ['018302', '1', '当前柜内温度'], ['018303', '1', '当前柜外温度']],
        68: [['068101', '2', '遥控电池组充电'], ['068102', '2', '遥控电池组放电'], ['068301', '1', '遥控电池组放电'],
             ['068302', '1', '电池组总电流'], ['068303', '1', '电池组01电流']],
        76: [['076301', '1', 'FSU硬盘占用率'], ['076302', '1', '内存占用率'], ['076303', '1', 'CPU使用率']],
        77: [['077101', '2', '是否开启风堵检测'], ['077102', '2', '是否开启风门检测'], ['077301', '1', '设备系统时间'],
             ['077302', '1', '直流电压'], ['077303', '1', '室内温度']],
        78: [['078101', '2', '模块01开机'], ['078102', '2', '模块01关机'], ['078301', '1', '相电压Ua'],
             ['078302', '1', '相电压Ub'], ['078303', '1', '相电压Uc']],
        87: [['087101', '2', '模块01开机'], ['087102', '2', '模块01关机'], ['087301', '1', '系统输入相电压Ua'],
             ['087302', '1', '系统输入相电压Ub'], ['087303', '1', '系统输入相电压Uc']],
        88: [['088301', '1', '直流电压'], ['088302', '1', '负载总电流01'], ['088303', '1', '分路01输出电流']],
        92: [['092301', '1', '相电压Ua'], ['092302', '1', '相电压Ub'], ['092303', '1', '相电压Uc']],
        17: [['017301', '1', '温度01'], ['017302', '1', '湿度01'], ['017303', '1', '水位01']],
        93: [['093301', '1', '分路01相电流Ia'], ['093302', '1', '分路01相电流Ib'], ['093303', '1', '分路01相电流Ic']]
    }

    return device, device_mete


def insert_fsu_data():
    """主函数：查询数据库并生成FSU、设备、信号三张表的数据"""
    device_types, device_mete = device_tuple()

    precinct_id = '01-01-08-07-05-01'
    device_id = '006011001768'
    device_id_fsu = 100100000000001

    precinct_id_sql = f"SELECT precinct_id FROM t_cfg_precinct WHERE precinct_id like '{precinct_id}%' and precinct_kind='5';"
    precinct_id_list = conn_obj.select(sql=precinct_id_sql)

    fsu_data_list = []
    device_data_list = []
    signal_data_list = []

    # 遍历每个precinct
    for room_idx, room_id in enumerate(precinct_id_list, 1):
        room_id_value = room_id[0]
        print(f"\n处理Precinct ({room_idx}/{len(precinct_id_list)}): {room_id_value}")

        # 查询device_code列表
        device_code_sql = f"select device_code from t_cfg_device where precinct_id = '{room_id_value}' and device_id like '%{device_id}%'; "
        device_code_list = conn_obj.select(sql=device_code_sql)

        # 为每个device_code生成数据
        for device_idx, device_item in enumerate(device_code_list, 1):
            fsuid = device_item[0]
            print(f"  处理FSU ({device_idx}/{len(device_code_list)}): {fsuid}")

            # 动态生成递增的站点和机房信息
            site_info = {
                'siteid': f'4401002000{room_idx:03d}',
                'sitename': f'测试站点{room_idx}',
                'roomid': f'4401002000{room_idx:03d}00',
                'roomname': f'测试机房{room_idx}'
            }

            # 1. 生成FSU数据
            fsu_row = {
                'fsuid': fsuid,
                'fsuname': f'测试FSU-{fsuid}',
                'fsuver': CONFIG['fsu_ver'],
                'siteid': site_info['siteid'],
                'sitename': site_info['sitename'],
                'roomid': site_info['roomid'],
                'roomname': site_info['roomname'],
                'interval': CONFIG['interval'],
                'm': CONFIG['m']
            }
            fsu_data_list.append(fsu_row)

            # 2. & 3. 生成设备和信号数据
            for devicetype, basename in device_types.items():
                device_id_str = str(device_id_fsu).zfill(15)

                # 生成设备数据（使用相同的site_info）
                device_row = {
                    'm': CONFIG['m'],
                    'fsuid': fsuid,
                    'deviceid': device_id_str,
                    'devicename': basename,
                    'devdescribe': '',
                    'siteid': site_info['siteid'],
                    'sitename': site_info['sitename'],
                    'roomid': site_info['roomid'],
                    'roomname': site_info['roomname'],
                    'devicetype': devicetype,
                    'devicesubtype': CONFIG['devicesubtype'],
                    'model': '',
                    'brand': CONFIG['brand'],
                    'ratedcapacity': CONFIG['ratedcapacity'],
                    'version': CONFIG['version'],
                    'beginruntime': '',
                    'confremark': ''
                }
                device_data_list.append(device_row)

                # 生成信号数据 - 根据配置生成多个通道号
                if devicetype in device_mete:
                    mete_list = device_mete[devicetype]

                    for mete_item in mete_list:
                        type_id, signal_type, signal_name = mete_item

                        # 为每个测点生成指定数量的通道号
                        for channel in range(CONFIG['signal_channels']):
                            signal_row = {
                                'm': CONFIG['m'],
                                'deviceid': device_id_str,
                                'Type': signal_type,
                                'ID': type_id,
                                'SignalName': signal_name,
                                'SignalNumber': str(channel).zfill(3),  # 格式化为三位数字：000, 001, 002...
                                'AlarmLevel': '2',
                                'Threshold': '0',
                                'NMAlarmID': ''
                            }
                            signal_data_list.append(signal_row)

                device_id_fsu += 1

    # 保存到Excel
    save_to_excel(fsu_data_list, device_data_list, signal_data_list)

    # 打印统计信息
    print(f"\n{'=' * 50}")
    print(f"数据生成完成！")
    print(f"FSU数量: {len(fsu_data_list)}")
    print(f"设备数量: {len(device_data_list)}")
    print(f"信号数量: {len(signal_data_list)}")
    print(f"文件已保存: fsu_data.xlsx")
    print(f"{'=' * 50}")


# ... save_to_excel 和 save_as_csv 函数保持不变 ...
def save_to_excel(fsu_data, device_data, signal_data):
    """使用openpyxl保存数据到Excel文件"""
    try:
        wb = Workbook()
        if wb.active:
            wb.remove(wb.active)

        # FSU表
        ws_fsu = wb.create_sheet('fsu')
        fsu_headers = ['fsuid', 'fsuname', 'fsuver', 'siteid', 'sitename', 'roomid', 'roomname', 'interval', 'm']
        ws_fsu.append(fsu_headers)
        for row in fsu_data:
            ws_fsu.append([row[h] for h in fsu_headers])

        # Device表
        ws_device = wb.create_sheet('device')
        device_headers = ['m', 'fsuid', 'deviceid', 'devicename', 'devdescribe', 'siteid', 'sitename',
                          'roomid', 'roomname', 'devicetype', 'devicesubtype', 'model', 'brand',
                          'ratedcapacity', 'version', 'beginruntime', 'confremark']
        ws_device.append(device_headers)
        for row in device_data:
            ws_device.append([row[h] for h in device_headers])

        # Signal表
        ws_signal = wb.create_sheet('signal')
        signal_headers = ['m', 'deviceid', 'Type', 'ID', 'SignalName', 'SignalNumber', 'AlarmLevel', 'Threshold',
                          'NMAlarmID']
        ws_signal.append(signal_headers)
        for row in signal_data:
            ws_signal.append([row[h] for h in signal_headers])

        wb.save('fsu_data.xlsx')
        print("Excel文件保存成功！")

    except Exception as e:
        print(f"Excel保存出错（可能是openpyxl未安装）: {e}")
        print("\n=== 自动切换到CSV格式 ===")
        save_as_csv(fsu_data, device_data, signal_data)


def save_as_csv(fsu_data, device_data, signal_data):
    """备选方案：保存为CSV文件"""
    try:
        # 保存FSU数据
        if fsu_data:
            with open('fsu.csv', 'w', encoding='utf-8-sig') as f:
                fsu_headers = ['fsuid', 'fsuname', 'fsuver', 'siteid', 'sitename', 'roomid', 'roomname', 'interval',
                               'm']
                f.write(','.join(fsu_headers) + '\n')
                for row in fsu_data:
                    f.write(','.join(str(row[h]) for h in fsu_headers) + '\n')
            print("已保存: fsu.csv")

        # 保存device数据
        if device_data:
            with open('device.csv', 'w', encoding='utf-8-sig') as f:
                device_headers = ['m', 'fsuid', 'deviceid', 'devicename', 'devdescribe', 'siteid', 'sitename',
                                  'roomid', 'roomname', 'devicetype', 'devicesubtype', 'model', 'brand',
                                  'ratedcapacity', 'version', 'beginruntime', 'confremark']
                f.write(','.join(device_headers) + '\n')
                for row in device_data:
                    f.write(','.join(str(row[h]) for h in device_headers) + '\n')
            print("已保存: device.csv")

        # 保存signal数据
        if signal_data:
            with open('signal.csv', 'w', encoding='utf-8-sig') as f:
                signal_headers = ['m', 'deviceid', 'Type', 'ID', 'SignalName', 'SignalNumber', 'AlarmLevel',
                                  'Threshold', 'NMAlarmID']
                f.write(','.join(signal_headers) + '\n')
                for row in signal_data:
                    f.write(','.join(str(row[h]) for h in signal_headers) + '\n')
            print("已保存: signal.csv")

        print("\nCSV文件已生成，可用Excel直接打开。")

    except Exception as csv_e:
        print(f"CSV保存也失败了: {csv_e}")


if __name__ == '__main__':
    try:
        insert_fsu_data()
    except NameError:
        print("错误: conn_obj 未定义！")
        print("请确保数据库连接对象 conn_obj 已经正确配置。")
    except Exception as e:
        print(f"执行出错: {e}")
