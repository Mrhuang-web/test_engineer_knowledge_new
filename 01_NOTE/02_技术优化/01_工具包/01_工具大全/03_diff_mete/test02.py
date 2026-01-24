import openpyxl


def compare_columns_refined(excel_file):
    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb['Sheet1']

    # 创建字典存储C列和M列的值及其对应的行索引
    c_dict = {}  # 键: C列值, 值: (行索引, D列值)
    m_dict = {}  # 键: M列值, 值: (行索引, N列值)

    # 读取C列和D列数据
    for row in range(2, sheet.max_row + 1):
        c_value = sheet.cell(row=row, column=3).value  # C列
        d_value = sheet.cell(row=row, column=4).value  # D列

        if c_value is not None:
            # 确保C列的值是字符串格式，保留前导零
            c_str = str(c_value).zfill(6) if isinstance(c_value, (int, float)) else str(c_value)
            c_dict[c_str] = (row, d_value)

    # 读取M列和N列数据
    for row in range(2, sheet.max_row + 1):
        m_value = sheet.cell(row=row, column=13).value  # M列
        n_value = sheet.cell(row=row, column=14).value  # N列

        if m_value is not None:
            # 确保M列的值是字符串格式，保留前导零
            m_str = str(m_value).zfill(6) if isinstance(m_value, (int, float)) else str(m_value)
            m_dict[m_str] = (row, n_value)

    # 找出C列有但M列没有的值
    c_only = set(c_dict.keys()) - set(m_dict.keys())

    # 找出M列有但C列没有的值
    m_only = set(m_dict.keys()) - set(c_dict.keys())

    # 找出C列和M列都有的值
    common = set(c_dict.keys()) & set(m_dict.keys())

    # 找出C列和M列相同但D列和N列不同的值
    common_but_different = []
    for common_value in common:
        c_row_index, d_value = c_dict[common_value]
        m_row_index, n_value = m_dict[common_value]

        # 比较D列和N列的值是否相同
        d_str = str(d_value) if d_value is not None else ""
        n_str = str(n_value) if n_value is not None else ""

        if d_str[:2] != n_str:
            common_but_different.append((common_value, d_value, n_value))

    # 添加标题
    sheet.cell(row=1, column=16).value = "标准字典表-t_cfg_mete仅有"  # O列标题
    sheet.cell(row=1, column=19).value = "现提供的excel仅有"  # Q列标题
    sheet.cell(row=1, column=22).value = "标准字典表-t_cfg_dict信号编码"  # V列标题
    sheet.cell(row=1, column=23).value = "信号量类型"  # W列标题
    sheet.cell(row=1, column=25).value = "现提供的excel信号编码"  # Y列标题
    sheet.cell(row=1, column=26).value = "信号量类型"  # Z列标题

    # 初始化写入位置
    o_row = 2  # O列起始行
    q_row = 2  # Q列起始行
    v_row = 2  # V列起始行

    # 将C列有但M列没有的值写入O列和P列
    for c_value in c_only:
        row_index, d_value = c_dict[c_value]
        sheet.cell(row=o_row, column=16).value = c_value  # O列
        sheet.cell(row=o_row, column=17).value = d_value  # P列
        o_row += 1

    # 将M列有但C列没有的值写入Q列和R列
    for m_value in m_only:
        row_index, n_value = m_dict[m_value]
        sheet.cell(row=q_row, column=19).value = m_value  # Q列
        sheet.cell(row=q_row, column=20).value = n_value  # R列
        q_row += 1

    # 将C列和M列相同但D列和N列不同的值写入V、W、Y、Z列
    for common_value, d_value, n_value in common_but_different:
        # 写入V列和W列（C列和D列数据）
        sheet.cell(row=v_row, column=22).value = common_value  # V列
        sheet.cell(row=v_row, column=23).value = d_value  # W列

        # 写入Y列和Z列（M列和N列数据）
        sheet.cell(row=v_row, column=25).value = common_value  # Y列
        sheet.cell(row=v_row, column=26).value = n_value  # Z列

        v_row += 1

    # 保存文件
    wb.save('./result/comparison_result_extended.xlsx')
    print("精炼比较完成，结果已保存到 comparison_result_refined.xlsx")

    # 打印统计信息
    print(f"C列有但M列没有的数据数量: {len(c_only)}")
    print(f"M列有但C列没有的数据数量: {len(m_only)}")
    print(f"C列和M列都有的数据数量: {len(common)}")
    print(f"C列和M列相同但D列和N列不同的数据数量: {len(common_but_different)}")


# 使用示例
compare_columns_refined('./excel/对比文件.xlsx')